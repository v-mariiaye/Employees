import pandas as pd
from openpyxl import Workbook

csv_file = 'employees.csv'
try:
    df = pd.read_csv(csv_file)
except FileNotFoundError:
    print("Помилка: Файл CSV не знайдено.")
    exit(1)
except Exception as e:
    print(f"Помилка: {e}. Неможливо відкрити файл CSV.")
    exit(1)

xlsx_file = 'employee_data.xlsx'
wb = Workbook()
ws_all = wb.active
ws_all.title = 'all'

column_names = list(df.columns)
ws_all.append(column_names)

for r_idx, row in enumerate(df.iterrows(), start=2):
    for c_idx, value in enumerate(row[1], start=1):
        ws_all.cell(row=r_idx, column=c_idx, value=value)

age_categories = {
    'younger_18': (0, 18),
    '18-45': (18, 45),
    '45-70': (45, 70),
    'older_70': (70, None)
}

for category, (min_age, max_age) in age_categories.items():
    ws = wb.create_sheet(title=category)
    ws.append(['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік'])
    r_idx = 2
    for _, row in df.iterrows():
        birth_date = pd.to_datetime(row['Дата народження'])
        age = (pd.Timestamp('now') - birth_date).days // 365
        if (max_age is None and age >= min_age) or (min_age <= age < max_age):
            data = [r_idx - 1, row['Прізвище'], row['Ім’я'], row['По батькові'], row['Дата народження'], age]
            ws.append(data)
            r_idx += 1

try:
    wb.save(xlsx_file)
    print(f"Ok: Файл {xlsx_file} успішно збережено.")
except Exception as e:
    print(f"Помилка: {e}. Неможливо зберегти файл {xlsx_file}.")